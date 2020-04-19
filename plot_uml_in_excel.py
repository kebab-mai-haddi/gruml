import logging
from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG)


class WriteInExcel:

    def __init__(self, file_name='UML_Spreadsheet.xlsx'):
        """[summary]

        Keyword Arguments:
            file_name {str} -- [description] (default: {'UML_Spreadsheet.xlsx'})
        """
        self.file_name = file_name
        self.writer = pd.ExcelWriter(self.file_name, engine='xlsxwriter')
        self.count = 0

    def get_number_of_rows_in_df(self, agg_data):
        """compute number of rows in the dataframe using aggregated data.

        Arguments:
            agg_data {defaultdict(list)} -- the complete data generated for the source code.

        Returns:
            int -- number of rows that will be plotted in the dataframe.
        """
        number_of_rows = 0
        for module in agg_data.keys():
            number_of_rows += 1
            for class_dict in agg_data[module]:
                number_of_rows += 1 + len(class_dict['Methods'])
        return number_of_rows

    def create_pandas_dataframe(self, agg_data, skip_cols):
        """create pandas dataframe to be plotted in the excel/csv sheet.

        Arguments:
            agg_data {list} -- the complete data generated for the source code.
            skip_cols {int} -- number of columns to be skipped on the left hand side of the 
                               class in the dataframe.

        Returns:
            pandas.DataFrame -- the final dataframe created which is to be further styled and plotted in excel.
        """
        number_of_rows = self.get_number_of_rows_in_df(agg_data)
        logging.debug('Number of rows in total: {}'.format(number_of_rows))
        columns = ['' for _ in range(skip_cols+2)]
        df = pd.DataFrame(
            index=np.arange(0, number_of_rows),
            columns=columns+['Module', 'Class/Functions', 'Methods']
        )
        row_counter = 0
        column_counter = skip_cols-1
        # mapping to store row, col for dependents.
        # parent: child mapping created just for plotting tree like line
        parent_to_child_mapping = defaultdict(lambda: defaultdict(list))
        dependee_to_dependents_mapping = defaultdict(lambda: defaultdict(list))
        # {module : {class: [row number, {methods: row_number}]}} mapping
        self.class_row_mapping = defaultdict(lambda: defaultdict(list))
        # mapping to store dark edges -> 0: [17, 10, 13] is a COLUMN:ROWS range.
        self.dark_edges_column = defaultdict(list)
        self.inheritance_edges_column = defaultdict(list)
        for module in agg_data.keys():
            logging.debug(
                "Currently framing df with module: {}".format(module))
            df.loc[row_counter] = columns + ['{}'.format(module), '', '']
            row_counter += 1
            for class_data in agg_data[module]:
                base_class = class_data['Class']
                self.class_row_mapping[module][base_class] = [row_counter, {}]
                methods = class_data['Methods']
                parents = class_data['Parents']
                dependents = class_data['Dependents']
                columns = ['' for _ in range(skip_cols+2)]
                df.loc[row_counter] = columns + ['', base_class, '']
                row_counter += 1
                for method in methods:
                    # logging.debug('Inserting method: {} of class: {} at row: {} and column: {}'.format(
                        # method, base_class, row_counter, skip_cols+1))
                    df.iloc[row_counter, skip_cols+4] = method
                    self.class_row_mapping[module][base_class][1][method] = (
                        row_counter, skip_cols+4)
                    row_counter += 1
                if parents:
                    for parent in parents:
                        parent_to_child_mapping[parent['parent_module']][parent['parent_class']].append(
                            {'child_module': module, 'child_class': base_class}
                        )
                if dependents:
                    for dependent in dependents:
                        dependee_to_dependents_mapping[module][base_class].append(
                            {'dependent_module': dependent['module'],
                                'dependent_class': dependent['class']}
                        )
        # get a single list of dependees and parents so as to draw a common vertical pipe
        dependees_and_parents_combined = defaultdict(set)
        for dependee_module in dependee_to_dependents_mapping.keys():
            for dependee in dependee_to_dependents_mapping[dependee_module]:
                dependees_and_parents_combined[dependee_module].add(dependee)
        for parent_module in parent_to_child_mapping.keys():
            for parent in parent_to_child_mapping[parent_module]:
                dependees_and_parents_combined[parent_module].add(parent)
        logging.debug("All classes that are parents or dependees are: ")
        logging.debug(dependees_and_parents_combined)
        for module in dependees_and_parents_combined:
            if not self.class_row_mapping.get(module, None):
                logging.warn(
                    "Out of scope module was brought up, no row maping found")
                continue
            for class_ in dependees_and_parents_combined[module]:
                if not self.class_row_mapping[module].get(class_, None):
                    logging.debug(
                        'This class is in dependees and parents combined but not in class row mapping: {}'.format(class_))
                    continue
                df.iloc[self.class_row_mapping[module]
                        [class_][0]][column_counter] = "⇒"
                self.dark_edges_column[column_counter].append(
                    self.class_row_mapping[module][class_][0]
                )
                if class_ in dependee_to_dependents_mapping[module]:
                    df.iloc[self.class_row_mapping[module]
                            [class_][0]][skip_cols] = "⇒"
                    for dependent in dependee_to_dependents_mapping[module][class_]:
                        df.iloc[self.class_row_mapping[dependent['dependent_module']][dependent['dependent_class']]
                                [0], column_counter] = "←"
                        self.dark_edges_column[column_counter].append(
                            self.class_row_mapping[dependent['dependent_module']][dependent['dependent_class']][0])
                if class_ in parent_to_child_mapping[module]:
                    df.iloc[self.class_row_mapping[module]
                            [class_][0]][skip_cols+1] = "▷"
                    for child in parent_to_child_mapping[module][class_]:
                        df.iloc[self.class_row_mapping[child['child_module']][child['child_class']]
                                [0], column_counter] = "◁"
                        self.dark_edges_column[column_counter].append(
                            self.class_row_mapping[child['child_module']][child['child_class']][0])
                column_counter -= 1
        # convert all NaN to None.
        df = df.replace(np.nan, '', regex=True)
        # logging.debug(
        #     "Create DF was called with skip columns as: {}".format(skip_cols))
        # logging.debug(df)
        return df

    def integrate_sequence_diagram_in_df(self, df, function_sequence, use_case, driver_function):
        """integrates sequence diagram in the dataframe.

        Arguments:
            df {pandas.DataFrame} -- dataframe that has to incorporate the sequence diagram.
            function_sequence {[type]} -- sequence of the functions called while tracing a particular sequence diagram.
            use_case {str} -- name of the use case for which we are drawing the sequence diagram.

        Returns:
            [type] -- 
        """
        # add new columns in dataframe
        number_of_columns_pre_sequence = len(df.columns)
        for event_number in range(len(function_sequence)):
            df['{}'.format(event_number)] = np.nan
        event_counter = 1
        # counter to check whether its the first or last column in sequence diagram section
        logging.debug("Class row mapping: ")
        logging.debug(self.class_row_mapping)
        logging.debug("Function sequence received to plot is: ")
        logging.debug(function_sequence)
        for event in function_sequence:
            logging.debug("Event is: {}".format(event))
            caller, callee = event
            logging.debug("caller is: {}".format(caller))
            if caller == driver_function:
                continue
            caller_class, caller_function = caller.split('.')
            logging.debug("Caller class is: {}, caller function is: {}".format(
                caller_class, caller_function))
            callee_class, callee_function = callee.split('.')
            logging.debug("Callee class is: {}, callee function is: {}".format(
                callee_class, callee_function
            ))
            if caller_class not in self.class_row_mapping or callee_class not in self.class_row_mapping:
                continue
            caller_row_number, caller_column_number = self.class_row_mapping[
                caller_class][1][caller_function]
            callee_row_number, callee_column_number = self.class_row_mapping[
                callee_class][1][callee_function]
            caller_column_number, callee_column_number = number_of_columns_pre_sequence + \
                event_counter - 1, number_of_columns_pre_sequence + event_counter - 1
            self.dark_edges_column[caller_column_number].extend(
                [caller_row_number, callee_row_number])
            if df.iloc[caller_row_number, caller_column_number] == "←":
                df.iloc[caller_row_number, caller_column_number] = "↔"
            else:
                df.iloc[caller_row_number, caller_column_number] = "→"
            if df.iloc[callee_row_number, callee_column_number] == "→":
                df.iloc[callee_row_number, callee_column_number] = "↔"
            else:
                df.iloc[callee_row_number, callee_column_number] = "←"
            event_counter += 1
        df = df.replace(np.nan, '', regex=True)
        return df

    def write_df_to_excel(self, df, sheet_name, skip_cols, classes={}, use_case=None):
        """write dataframe to excel

        Arguments:
            df {pandas.DataFrame} -- the dataframe that we have to write to excel after making some changes.
            sheet_name {str} -- name of the target excel sheet/
            skip_cols {int} -- number of columns to be skipped for drawing out arrows.

        Keyword Arguments:
            classes {dict} -- name of the classes (default: {{}})
            use_case {str} -- use case for which we are plotting a particular sequence diagram (default: {None})
        """
        # logging.debug("Use Case as an argument is: {}".format(use_case))
        self.count += 1
        if self.count == 2:
            self.file_name = 'Use_Case_{}'.format(use_case) + self.file_name
            self.writer = pd.ExcelWriter(self.file_name, engine='xlsxwriter')
        df.to_excel(self.writer, sheet_name=sheet_name,
                    header=True, index=False)
        self.writer.save()
        self.writer.close()
        import sys
        wb = load_workbook(filename=self.file_name)
        ws = wb[sheet_name]
        ws.column_dimensions['{}'.format(get_column_letter(
            skip_cols+1))].width = 8.4  # len(Parent) = (6+1)*1.2
        # len(Methods/Children) + 1(M) + 1(C) + 1(/) = 19*1.2
        ws.column_dimensions['{}'.format(
            get_column_letter(skip_cols+2))].width = 22.8
        bd = Side(style='thick', color='000000')
        # to check whether a col in sheet's columns has arrived for dark edges.
        col_check_counter = 0
        logging.debug('Skip columns are: {}'.format(skip_cols))
        for _ in ws.columns:
            if col_check_counter in self.dark_edges_column:
                logging.debug(
                    'Col {} is in dark edges column'.format(col_check_counter))
                column_letter = get_column_letter(col_check_counter+1)
                row_range = sorted(self.dark_edges_column[col_check_counter])
                '''
                plus one because dataframe's 0 is excel'1 but column headings are on 1 so we need to start from 2.
                the second addition of changing 0-indexed to 1-indexed comes in the next line while plotting.
                Now, second index of list is +2 so that we can cover the full range(0 to 3 is to be covered,
                so for loop would be from range(0,4))
                get the first and last row to draw the dark edges.
                '''
                for row_iterator in range(row_range[0]+1, row_range[-1]+2):
                    logging.debug('{}{}'.format(column_letter, row_iterator+1))
                    if col_check_counter >= skip_cols:  # should it be > or >=
                        ws['{}{}'.format(column_letter, row_iterator+1)
                           ].border = Border(right=bd)
                    else:
                        ws['{}{}'.format(column_letter, row_iterator+1)
                           ].border = Border(left=bd)
            col_check_counter += 1

        red_fill = PatternFill(
            start_color='FFFF0000',
            end_color='FFFF0000',
            fill_type='solid'
        )
        for row in range(1, ws.max_row+1):
            column_letter = get_column_letter(skip_cols+1)
            if ws['{}{}'.format(column_letter, row)].value:
                ws['{}{}'.format(column_letter, row)].fill = red_fill
            column_letter = get_column_letter(skip_cols+2)
            if ws['{}{}'.format(column_letter, row)].value:
                ws['{}{}'.format(column_letter, row)].fill = red_fill
        # add a new row in the worksheet
        if use_case:
            logging.debug("Working for the Use Case: {}".format(use_case))
            ws.insert_rows(0)
            use_case_column_letter = get_column_letter(skip_cols+4)
            use_case_cell = '{}{}'.format(use_case_column_letter, 1)
            logging.debug("Use Case cell is {}".format(use_case_cell))
            ws[use_case_cell] = use_case
        # adjust the width of all columns
        col_counter = 1
        for _ in ws.columns:
            if col_counter > skip_cols:
                if col_counter == skip_cols+3:
                    ws.column_dimensions[get_column_letter(
                        col_counter)].width = 6
                elif col_counter > skip_cols+3:
                    ws.column_dimensions[get_column_letter(
                        col_counter)].width = 24
                else:
                    ws.column_dimensions[get_column_letter(
                        col_counter)].width = 3
                col_counter += 1
                continue
            else:
                ws.column_dimensions[get_column_letter(
                    col_counter)].hidden = True
            ws.column_dimensions[get_column_letter(col_counter)].width = 3
            col_counter += 1
        # give bold font to cells with Classes, to classless functions, and, to all modules
        modules_column = get_column_letter(skip_cols+3)
        classes_column = get_column_letter(skip_cols+4)
        font = Font(bold=True)
        values_to_be_bold = set()
        for module in classes.keys():
            values_to_be_bold.add('{}'.format(module))
            for class_ in classes[module]:
                values_to_be_bold.add(class_)
        for row in range(2, ws.max_row+1):
            if ws['{}{}'.format(classes_column, row)].value:
                ws['{}{}'.format(classes_column, row)].font = font
            elif ws['{}{}'.format(modules_column, row)].value:
                ws['{}{}'.format(modules_column, row)].font = font
            else:
                logging.debug("Hiding this row: {}".format(row))
                ws.row_dimensions[row].hidden = True
        # show the Class row when use-cases are there.
        if use_case:
            ws.row_dimensions[2].hidden = False
        wb.save(self.file_name)
        logging.debug("{}:{} done!".format(self.file_name, sheet_name))
