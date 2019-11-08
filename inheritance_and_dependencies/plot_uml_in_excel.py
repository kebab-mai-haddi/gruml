from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


class WriteInExcel:

    def __init__(self, file_name='UML_Spreadsheet.xlsx'):
        self.file_name = file_name
        self.writer = pd.ExcelWriter(self.file_name, engine='xlsxwriter')

    def get_number_of_rows_in_df(self, agg_data):
        number_of_rows = 0
        for class_dict in agg_data:
            number_of_rows += 1 + \
                len(class_dict['Methods']) + len(class_dict['Children'])
        return number_of_rows

    def create_pandas_dataframe(self, agg_data, skip_cols):
        number_of_rows = self.get_number_of_rows_in_df(agg_data)
        print('Number of rows in total: {}'.format(number_of_rows))
        columns = ['' for _ in range(skip_cols)]
        df = pd.DataFrame(
            index=np.arange(0, number_of_rows),
            columns=columns+['Parent', 'Methods/Children']
        )
        row_counter = 0
        column_counter = skip_cols-1
        # mapping to store row, col for dependents.
        dependents_col_counter = {}
        class_row_mapping = {}  # class: row number mapping
        self.dark_edges_column = defaultdict(list)
        prev_class_row_counter = 0
        for class_data in agg_data:
            base_class = class_data['Class']
            class_row_mapping[base_class] = row_counter
            methods = class_data['Methods']
            children = class_data['Children']
            dependents = class_data['Dependents']
            columns = ['' for _ in range(skip_cols)]
            df.loc[row_counter] = columns + [base_class, '']
            row_counter += 1
            for method in methods:
                print('Inserting method: {} of class: {} at row: {} and column: {}'.format(
                    method, base_class, row_counter, skip_cols+1))
                df.iloc[row_counter, skip_cols+1] = method
                row_counter += 1
            for child in children:
                df.iloc[row_counter, skip_cols+1] = child
                row_counter += 1
            for dependent in dependents:
                df.iloc[prev_class_row_counter, column_counter] = "→"
                self.dark_edges_column[column_counter].append(
                    prev_class_row_counter)
                if dependents_col_counter.get(dependent, None):
                    dependents_col_counter[dependent].append(column_counter)
                else:
                    dependents_col_counter[dependent] = [column_counter]
                column_counter -= 1
            prev_class_row_counter = row_counter

        for dependent in dependents_col_counter.keys():
            for column in dependents_col_counter[dependent]:
                df.iloc[class_row_mapping[dependent], column] = "←"
                self.dark_edges_column[column].append(
                    class_row_mapping[dependent])

        # convert all NaN to None.
        df = df.replace(np.nan, '', regex=True)
        print("Create DF was called.")
        print(df)
        return df

    def write_df_to_excel(self, df, sheet_name):
        df.to_excel(self.writer, sheet_name=sheet_name,
                    header=True, index=False)
        self.writer.save()
        self.writer.close()
        wb = load_workbook(filename=self.file_name)
        ws = wb[sheet_name]
        bd = Side(style='thick', color='000000')
        # to check whether a col in sheet's columns has arrived for dark edges.
        col_check_counter = 0
        for _ in ws.columns:
            if col_check_counter in self.dark_edges_column:
                column_letter = get_column_letter(col_check_counter+1)
                row_range = sorted(self.dark_edges_column[col_check_counter])
                # plus one because dataframe's 0 is excel'1 but column headings are on 1 so we need to start from 2.
                # the second addition of changing 0-indexed to 1-indexed comes in the next line while plotting.
                # Now, second index of list is +2 so that we can cover the full range(0 to 3 is to be covered,
                # so for loop would be from range(0,4))
                for row_iterator in range(row_range[0]+1, row_range[1]+2):
                    ws['{}{}'.format(column_letter, row_iterator+1)
                       ].border = Border(left=bd)
            col_check_counter += 1
        wb.save(self.file_name)
        print("{} done!".format(sheet_name))
